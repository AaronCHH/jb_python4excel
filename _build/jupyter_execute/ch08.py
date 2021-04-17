# Ch08 Excel File Manipulation with Reader and Writer 

## The Reader and Writer Packages

### OpenPyXL

### Reading with OpenPyXL

import pandas as pd
import openpyxl
import excel
import datetime as dt

# Open the workbook to read cell values.
# The file is automatically closed again after loading the data.
book = openpyxl.load_workbook("xl/stores.xlsx", data_only=True)

# Get a worksheet object by name or index (0-based)
sheet = book["2019"]
sheet = book.worksheets[0]

# Get a list with all sheet names
book.sheetnames

# Loop through the sheet objects.
# Instead of "name", openpyxl uses "title".
for i in book.worksheets:
    print(i.title)

# Getting the dimensions,
# i.e. the used range of the sheet
sheet.max_row, sheet.max_column

# Read the value of a single cell
# using "A1" notation and using cell indices (1-based)
sheet["B6"].value
sheet.cell(row=6, column=2).value

# Read in a range of cell values by using our excel module
data = excel.read(book["2019"], (2, 2), (8, 6))
data[:2]  # Print the first two rows

### Writing with OpenPyXL

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill
import excel

# Instantiate a workbook
book = openpyxl.Workbook()

# Get the first sheet and give it a name
sheet = book.active
sheet.title = "Sheet1"

# Writing individual cells using A1 notation
# and cell indices (1-based)
sheet["A1"].value = "Hello 1"
sheet.cell(row=2, column=1, value="Hello 2")

# Formatting: fill color, alignment, border and font
font_format = Font(color="FF0000", bold=True)
thin = Side(border_style="thin", color="FF0000")
sheet["A3"].value = "Hello 3"
sheet["A3"].font = font_format
sheet["A3"].border = Border(top=thin, left=thin,
                            right=thin, bottom=thin)
sheet["A3"].alignment = Alignment(horizontal="center")
sheet["A3"].fill = PatternFill(fgColor="FFFF00", fill_type="solid")

# Number formatting (using Excel's formatting strings)
sheet["A4"].value = 3.3333
sheet["A4"].number_format = "0.00"

# Date formatting (using Excel's formatting strings)
sheet["A5"].value = dt.date(2016, 10, 13)
sheet["A5"].number_format = "mm/dd/yy"

# Formula: you must use the English name of the formula
# with commas as delimiters
sheet["A6"].value = "=SUM(A4, 2)"

# Image
sheet.add_image(Image("images/python.png"), "C1")

# Two-dimensional list (we're using our excel module)
data = [[None, "North", "South"],
        ["Last Year", 2, 5],
        ["This Year", 3, 6]]
excel.write(sheet, data, "A10")

# Chart
chart = BarChart()
chart.type = "col"
chart.title = "Sales Per Region"
chart.x_axis.title = "Regions"
chart.y_axis.title = "Sales"
chart_data = Reference(sheet, min_row=11, min_col=1,
                       max_row=12, max_col=3)
chart_categories = Reference(sheet, min_row=10, min_col=2,
                             max_row=10, max_col=3)
# from_rows interprets the data in the same way
# as if you would add a chart manually in Excel
chart.add_data(chart_data, titles_from_data=True, from_rows=True)
chart.set_categories(chart_categories)
sheet.add_chart(chart, "A15")

# Saving the workbook creates the file on disk
book.save("openpyxl.xlsx")

book = openpyxl.Workbook()
sheet = book.active
sheet["A1"].value = "This is a template"
book.template = True
book.save("template.xltx")

### Editing with OpenPyXL

# Read the stores.xlsx file, change a cell
# and store it under a new location/name.
book = openpyxl.load_workbook("xl/stores.xlsx")
book["2019"]["A1"].value = "modified"
book.save("stores_edited.xlsx")

book = openpyxl.load_workbook("xl/macro.xlsm", keep_vba=True)
book["Sheet1"]["A1"].value = "Click the button!"
book.save("macro_openpyxl.xlsm")

### XlsxWriter

import datetime as dt
import xlsxwriter
import excel

# Instantiate a workbook
book = xlsxwriter.Workbook("xlxswriter.xlsx")

# Add a sheet and give it a name
sheet = book.add_worksheet("Sheet1")

# Writing individual cells using A1 notation
# and cell indices (0-based)
sheet.write("A1", "Hello 1")
sheet.write(1, 0, "Hello 2")

# Formatting: fill color, alignment, border and font
formatting = book.add_format({"font_color": "#FF0000",
                              "bg_color": "#FFFF00",
                              "bold": True, "align": "center",
                              "border": 1, "border_color": "#FF0000"})
sheet.write("A3", "Hello 3", formatting)

# Number formatting (using Excel's formatting strings)
number_format = book.add_format({"num_format": "0.00"})
sheet.write("A4", 3.3333, number_format)

# Date formatting (using Excel's formatting strings)
date_format = book.add_format({"num_format": "mm/dd/yy"})
sheet.write("A5", dt.date(2016, 10, 13), date_format)

# Formula: you must use the English name of the formula
# with commas as delimiters
sheet.write("A6", "=SUM(A4, 2)")

# Image
sheet.insert_image(0, 2, "images/python.png")

# Two-dimensional list (we're using our excel module)
data = [[None, "North", "South"],
        ["Last Year", 2, 5],
        ["This Year", 3, 6]]
excel.write(sheet, data, "A10")

# Chart: see the file "sales_report_xlsxwriter.py" in the
# companion repo to see how you can work with indices
# instead of cell addresses
chart = book.add_chart({"type": "column"})
chart.set_title({"name": "Sales per Region"})
chart.add_series({"name": "=Sheet1!A11",
                  "categories": "=Sheet1!B10:C10",
                  "values": "=Sheet1!B11:C11"})
chart.add_series({"name": "=Sheet1!A12",
                  "categories": "=Sheet1!B10:C10",
                  "values": "=Sheet1!B12:C12"})
chart.set_x_axis({"name": "Regions"})
chart.set_y_axis({"name": "Sales"})
sheet.insert_chart("A15", chart)

# Closing the workbook creates the file on disk
book.close()

### book = xlsxwriter.Workbook("macro_xlxswriter.xlsm")
sheet = book.add_worksheet("Sheet1")
sheet.write("A1", "Click the button!")
book.add_vba_project("xl/vbaProject.bin")
sheet.insert_button("A3", {"macro": "Hello", "caption": "Button 1",
                           "width": 130, "height": 35})
book.close()

### pyxlsb

import pyxlsb
import excel

# Loop through sheets. With pyxlsb, the workbook
# and sheet objects can be used as context managers.
# book.sheets returns a list of sheet names, not objects!
# To get a sheet object, use get_sheet() instead.
with pyxlsb.open_workbook("xl/stores.xlsb") as book:
    for sheet_name in book.sheets:
        with book.get_sheet(sheet_name) as sheet:
            dim = sheet.dimension
            print(f"Sheet '{sheet_name}' has " 
                  f"{dim.h} rows and {dim.w} cols")

# Read in the values of a range of cells by using our excel module.
# Instead of "2019", you could also use its index (1-based).
with pyxlsb.open_workbook("xl/stores.xlsb") as book:
    with book.get_sheet("2019") as sheet:
        data = excel.read(sheet, "B2")
data[:2]  # Print the first two rows

from pyxlsb import convert_date
convert_date(data[1][3])

df = pd.read_excel("xl/stores.xlsb", engine="pyxlsb")

### xlrd, xlwt and xlutils

### Reading with xlrd

import xlrd
import xlwt
from xlwt.Utils import cell_to_rowcol2
import xlutils
import excel

# Open the workbook to read cell values. The file is
# automatically closed again after loading the data.
book = xlrd.open_workbook("xl/stores.xls")

# Get a list with all sheet names
book.sheet_names()

# Loop through the sheet objects
for sheet in book.sheets():
    print(sheet.name)

# Get a sheet object by name or index (0-based)
sheet = book.sheet_by_index(0)
sheet = book.sheet_by_name("2019")

# Dimensions
sheet.nrows, sheet.ncols

# Read the value of a single cell
# using "A1" notation and using cell indices (0-based).
# The "*" unpacks the tuple that cell_to_rowcol2 returns 
# into individual arguments.
sheet.cell(*cell_to_rowcol2("B3")).value
sheet.cell(2, 1).value

# Read in a range of cell values by using our excel module
data = excel.read(sheet, "B2")
data[:2]  # Print the first two rows

### Writing with xlwt

import xlwt
from xlwt.Utils import cell_to_rowcol2
import datetime as dt
import excel

# Instantiate a workbook
book = xlwt.Workbook()

# Add a sheet and give it a name
sheet = book.add_sheet("Sheet1")

# Writing individual cells using A1 notation
# and cell indices (0-based)
sheet.write(*cell_to_rowcol2("A1"), "Hello 1")
sheet.write(r=1, c=0, label="Hello 2")

# Formatting: fill color, alignment, border and font
formatting = xlwt.easyxf("font: bold on, color red;"
                         "align: horiz center;"
                         "borders: top_color red, bottom_color red,"
                                  "right_color red, left_color red,"
                                  "left thin, right thin,"
                                  "top thin, bottom thin;"
                         "pattern: pattern solid, fore_color yellow;")
sheet.write(r=2, c=0, label="Hello 3", style=formatting)

# Number formatting (using Excel's formatting strings)
number_format = xlwt.easyxf(num_format_str="0.00")
sheet.write(3, 0, 3.3333, number_format)

# Date formatting (using Excel's formatting strings)
date_format = xlwt.easyxf(num_format_str="mm/dd/yyyy")
sheet.write(4, 0, dt.datetime(2012, 2, 3), date_format)

# Formula: you must use the English name of the formula
# with commas as delimiters
sheet.write(5, 0, xlwt.Formula("SUM(A4, 2)"))

# Two-dimensional list (we're using our excel module)
data = [[None, "North", "South"],
        ["Last Year", 2, 5],
        ["This Year", 3, 6]]
excel.write(sheet, data, "A10")

# Picture (only allows to add bmp format)
sheet.insert_bitmap("images/python.bmp", 0, 2)

# This writes the file to disk
book.save("xlwt.xls")

### Editing with xlutils

import xlutils.copy

book = xlrd.open_workbook("xl/stores.xls", formatting_info=True)
book = xlutils.copy.copy(book)
book.get_sheet(0).write(0, 0, "changed!")
book.save("stores_edited.xls")

# Advanced Topics
## Working with Big Files

### Writing with OpenPyXL

book = openpyxl.Workbook(write_only=True)
# With write_only=True, book.active doesn't work
sheet = book.create_sheet()
# This will produce a sheet with 1000 x 200 cells
for row in range(1000):
    sheet.append(list(range(200)))
book.save("openpyxl_optimized.xlsx")

### Writing with XlsxWriter

book = xlsxwriter.Workbook("xlsxwriter_optimized.xlsx",
                           options={"constant_memory": True})
sheet = book.add_worksheet()
# This will produce a sheet with 1000 x 200 cells
for row in range(1000):
    sheet.write_row(row , 0, list(range(200)))
book.close()

### Reading with xlrd

with xlrd.open_workbook("xl/stores.xls", on_demand=True) as book:
    sheet = book.sheet_by_index(0)  # Only loads the first sheet

with xlrd.open_workbook("xl/stores.xls", on_demand=True) as book:
    with pd.ExcelFile(book, engine="xlrd") as f:
        df = pd.read_excel(f, sheet_name=0)

### Reading with OpenPyXL

book = openpyxl.load_workbook("xl/big.xlsx",
                              data_only=True, read_only=True,
                              keep_links=False)
# Perform the desired read operations here
book.close()  # Required with read_only=True

### Reading in Parallel

%%time
data = pd.read_excel("xl/big.xlsx",
                     sheet_name=None, engine="openpyxl")

%%time
import parallel_pandas
data = parallel_pandas.read_excel("xl/big.xlsx", sheet_name=None)

## Formatting DataFrames in Excel

with pd.ExcelFile("xl/stores.xlsx", engine="openpyxl") as xlfile:
    # Read a DataFrame
    df = pd.read_excel(xlfile, sheet_name="2020")

    # Get the OpenPyXL workbook object
    book = xlfile.book

    # From here on, it's OpenPyXL code
    sheet = book["2019"]
    value = sheet["B3"].value  # Read a single value

with pd.ExcelWriter("pandas_and_openpyxl.xlsx",
                    engine="openpyxl") as writer:
    df = pd.DataFrame({"col1": [1, 2, 3, 4], "col2": [5, 6, 7, 8]})
    # Write a DataFrame
    df.to_excel(writer, "Sheet1", startrow=4, startcol=2)

    # Get the OpenPyXL workbook and sheet objects
    book = writer.book
    sheet = writer.sheets["Sheet1"]

    # From here on, it's OpenPyXL code
    sheet["A1"].value = "This is a Title"  # Write a single cell value

df = pd.DataFrame({"col1": [1, -2], "col2": [-3, 4]},
                   index=["row1", "row2"])
df.index.name = "ix"
df

from openpyxl.styles import PatternFill

with pd.ExcelWriter("formatting_openpyxl.xlsx",
                    engine="openpyxl") as writer:
    # Write out the df with the default formatting to A1
    df.to_excel(writer, startrow=0, startcol=0)

    # Write out the df with custom index/header formatting to A6    
    startrow, startcol = 0, 5
    # 1. Write out the data part of the DataFrame
    df.to_excel(writer, header=False, index=False,
                startrow=startrow + 1, startcol=startcol + 1)
    # Get the sheet object and create a style object
    sheet = writer.sheets["Sheet1"]
    style = PatternFill(fgColor="D9D9D9", fill_type="solid")

    # 2. Write out the styled column headers
    for i, col in enumerate(df.columns):
        sheet.cell(row=startrow + 1, column=i + startcol + 2,
                   value=col).fill = style

    # 3. Write out the styled index
    index = [df.index.name if df.index.name else None] + list(df.index)
    for i, row in enumerate(index):
        sheet.cell(row=i + startrow + 1, column=startcol + 1,
                   value=row).fill = style

# Formatting index/headers with XlsxWriter
with pd.ExcelWriter("formatting_xlsxwriter.xlsx",
                    engine="xlsxwriter") as writer:
    # Write out the df with the default formatting to A1
    df.to_excel(writer, startrow=0, startcol=0)

    # Write out the df with custom index/header formatting to A6
    startrow, startcol = 0, 5
    # 1. Write out the data part of the DataFrame
    df.to_excel(writer, header=False, index=False,
                startrow=startrow + 1, startcol=startcol + 1)
    # Get the book and sheet object and create a style object
    book = writer.book
    sheet = writer.sheets["Sheet1"]
    style = book.add_format({"bg_color": "#D9D9D9"})

    # 2. Write out the styled column headers
    for i, col in enumerate(df.columns):
        sheet.write(startrow, startcol + i + 1, col, style)

    # 3. Write out the styled index
    index = [df.index.name if df.index.name else None] + list(df.index)
    for i, row in enumerate(index):
        sheet.write(startrow + i, startcol, row, style)

from openpyxl.styles import Alignment

with pd.ExcelWriter("data_format_openpyxl.xlsx",
                    engine="openpyxl") as writer:
    # Write out the DataFrame
    df.to_excel(writer)
    
    # Get the book and sheet objects
    book = writer.book
    sheet = writer.sheets["Sheet1"]
    
    # Formatting individual cells
    nrows, ncols = df.shape
    for row in range(nrows):
        for col in range(ncols):
            # +1 to account for the header/index
            # +1 since OpenPyXL is 1-based
            cell = sheet.cell(row=row + 2,
                              column=col + 2)
            cell.number_format = "0.000"
            cell.alignment = Alignment(horizontal="center")

with pd.ExcelWriter("data_format_xlsxwriter.xlsx",
                    engine="xlsxwriter") as writer:
    # Write out the DataFrame
    df.to_excel(writer)

    # Get the book and sheet objects    
    book = writer.book
    sheet = writer.sheets["Sheet1"]
    
    # Formatting the columns (individual cells can't be formatted)
    number_format = book.add_format({"num_format": "0.000",
                                     "align": "center"})
    sheet.set_column(first_col=1, last_col=2,
                     cell_format=number_format)

df.style.applymap(lambda x: "number-format: 0.000;"
                            "text-align: center")\
        .to_excel("styled.xlsx")

df = pd.DataFrame({"Date": [dt.date(2020, 1, 1)],
                   "Datetime": [dt.datetime(2020, 1, 1, 10)]})
with pd.ExcelWriter("date.xlsx",
                    date_format="yyyy-mm-dd",
                    datetime_format="yyyy-mm-dd hh:mm:ss") as writer:
    df.to_excel(writer)

